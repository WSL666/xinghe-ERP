from __future__ import annotations

import json
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

from ._base import MAX_PARALLEL, OUTPUT_DIR, PipelineStepError, VISION_TIMEOUT, call_text_llm, load_prompt_module, log, parse_json_response, require_env
from .generation import build_edit_image, generate_one_image
from .images import collect_product_images, summarize_image_inputs
from .vision import analyze_product_with_retry


def step1_read_xlsx(xlsx_path: str) -> list[dict]:
    """Read xlsx, extract product title and carousel images."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col

    title_col = None
    carousel_col = None
    for key, col in headers.items():
        if "产品标题" in key or key == "产品标题":
            title_col = col
        if "轮播图" in key or key == "轮播图":
            carousel_col = col

    if title_col is None:
        title_col = 1
    if carousel_col is None:
        carousel_col = 20

    results = []
    for row_idx in range(2, ws.max_row + 1):
        title = ws.cell(row=row_idx, column=title_col).value
        carousel_raw = ws.cell(row=row_idx, column=carousel_col).value
        if not title or not str(title).strip():
            continue

        carousel_images = []
        if carousel_raw:
            carousel_text = str(carousel_raw).strip()
            urls = re.split(r'[\n\r]+', carousel_text)
            carousel_images = [u.strip() for u in urls if u.strip().startswith("http")]
            if len(carousel_images) > 10:
                carousel_images = carousel_images[:10]

        results.append({
            "row": row_idx,
            "chinese_title": str(title).strip(),
            "carousel_images": carousel_images,
        })

    wb.close()
    return results


def step2_translate_titles(env: dict[str, str], products: list[dict]) -> list[dict]:
    """DeepSeek batch title translation."""
    log("=" * 50)
    log(">>> STEP2: DeepSeek title translation start")
    prompt_template = load_prompt_module("step2_translate.py")
    titles = [p["chinese_title"] for p in products]
    titles_json = json.dumps(titles, ensure_ascii=False)

    full_prompt = f"{prompt_template}\n\nInput:\n{titles_json}"

    step2_base_url = env.get("step2_base_url", "").strip() or None
    step2_api_key = env.get("step2_api_key", "").strip() or None
    step2_model = env.get("step2_model", "").strip() or None

    log(f"translate model: {step2_model}  endpoint: {step2_base_url}")
    log(f"original title: {titles[0] if titles else 'N/A'}")

    raw_text = call_text_llm(
        env, full_prompt, max_tokens=4096,
        base_url=step2_base_url, api_key=step2_api_key, model=step2_model,
    )
    log(f"translate raw response (first 300 chars): {raw_text[:300]}")
    translated = parse_json_response(raw_text)

    if not isinstance(translated, list):
        raise ValueError(f"translation result is not an array: {raw_text[:300]}")

    results = []
    for i, item in enumerate(translated):
        results.append({
            "row": products[i]["row"] if i < len(products) else i + 2,
            "chinese_title": item.get("cn_title") or item.get("chinese_title",
                                    products[i]["chinese_title"] if i < len(products) else ""),
            "english_title": item.get("en_title") or item.get("english_title", ""),
            "carousel_images": products[i]["carousel_images"] if i < len(products) else [],
        })
    log(f"translate done: cn={results[0]['chinese_title'][:80]}...")
    log(f"translate done: en={results[0]['english_title'][:80]}...")
    log("<<< STEP2 done")
    return results
    log("<<< STEP2 done")
    return results


def step3_analyze_vision(
    env: dict[str, str],
    products: list[dict],
    progress_callback=None,
) -> dict:
    """Step3: download images, call Vision, validate output JSON."""
    from prompts.vision_prompt import build_prompt as build_vision_prompt

    log("=" * 50)
    log(">>> STEP3: Vision analysis start")

    image_context = collect_product_images(products)
    product = products[0] if products else {}
    old_image_urls = product.get("old_image_urls", [])
    # vision consumes the raw chinese title directly; step2 translation
    # runs in parallel and is intentionally not awaited, so we never
    # block vision waiting for it (saves wall-clock time).
    product_text = product.get("chinese_title", "")
    vision_prompt = build_vision_prompt(product_text)

    log(f"Vision prompt product_text: {product_text[:120]}")
    log(f">>> calling Vision model ({len(image_context['valid_b64'])} images), timeout {VISION_TIMEOUT}s...")
    t0 = time.perf_counter()

    analysis = analyze_product_with_retry(
        env,
        vision_prompt,
        image_context["valid_b64"],
        image_context["valid_images"],
    )
    payload = analysis["payload"]
    selected_indexes = analysis["selected_indexes"]
    prompt_items = analysis["prompt_items"]

    vision_elapsed = time.perf_counter() - t0
    log(f"Vision done ({vision_elapsed:.1f}s): selected_indexes={selected_indexes}, "
        f"prompt_count={len(prompt_items)}")
    for num, pt in prompt_items:
        log(f"  image_{num}: {pt[:100]}...")

    result = {
        "step": "step3_vision",
        "en_title": product_text,
        "vision_prompt": vision_prompt,
        "payload": payload,
        "selected_indexes": selected_indexes,
        "prompt_items": [{"number": n, "prompt": p} for n, p in prompt_items],
        "image_context": {
            **summarize_image_inputs(image_context),
            "old_image_urls": old_image_urls,
        },
        "attempts": analysis["attempts"],
        "elapsed": vision_elapsed,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    meta_path = OUTPUT_DIR / f"step3_vision_{run_ts}.json"
    meta_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    result["meta_path"] = str(meta_path)
    # In-memory image cache for step4 reuse. Not persisted to the meta
    # json (bytes are not JSON-serializable); step4 falls back to a fresh
    # download when the cache is absent (e.g. manual run from the db).
    result["_image_cache"] = {
        "image_bytes_list": image_context["image_bytes_list"],
        "valid_indices": image_context["valid_indices"],
    }

    log("<<< STEP3 Vision done")
    return result
    log("<<< STEP3 Vision done")
    return result


def step4_generate_images(
    env: dict[str, str],
    products: list[dict],
    vision_result: dict,
    progress_callback=None,
) -> dict:
    """Step4: run I2I generation based on Step3 Vision result."""
    import traceback as _tb

    log("=" * 50)
    log(">>> STEP4: I2I generation start")

    vibe_api_key = require_env(env, "VIBE_API_KEY")
    vibe_base_url = require_env(env, "VIBE_BASE_URL")
    image_model = env.get("IMAGE_MODEL", "gpt-image-2")
    size = env.get("IMAGE_SIZE", "1024x1024")

    product = products[0] if products else {}
    old_image_urls = product.get("old_image_urls", [])
    selected_indexes = [int(idx) for idx in vision_result.get("selected_indexes", [])]
    raw_prompt_items = vision_result.get("prompt_items", [])

    # Reuse images already downloaded by step3 when available (the auto
    # pipeline hands the vision result over in-memory). Fall back to a
    # fresh download for manual runs where vision_result came from the
    # database without the cache.
    cache = vision_result.get("_image_cache") if isinstance(vision_result, dict) else None
    if cache and cache.get("image_bytes_list") is not None and cache.get("valid_indices") is not None:
        log(f"reuse step3 downloaded images: {len(cache['valid_indices'])}/{len(cache['image_bytes_list'])} usable")
        image_context = {
            "total_input_images": len(cache["image_bytes_list"]),
            "valid_images": len(cache["valid_indices"]),
            "valid_indices": cache["valid_indices"],
            "image_bytes_list": cache["image_bytes_list"],
        }
    else:
        image_context = collect_product_images(products)
        log(f"downloaded images fresh: {len(image_context['valid_indices'])}/{len(image_context['image_bytes_list'])} usable")

    prompt_items: list[tuple[int, str]] = []
    for item in raw_prompt_items:
        if isinstance(item, dict):
            prompt_items.append((int(item["number"]), str(item["prompt"])))
        else:
            prompt_items.append((int(item[0]), str(item[1])))

    if not selected_indexes:
        raise PipelineStepError("Step4 missing Vision selected reference image indexes", {
            "selected_indexes": selected_indexes,
            "prompt_count": len(prompt_items),
        })
    if not prompt_items:
        raise PipelineStepError("Step4 missing Vision generation prompts", {
            "selected_indexes": selected_indexes,
            "prompt_count": 0,
        })

    valid_indices = image_context["valid_indices"]
    image_bytes_list = image_context["image_bytes_list"]
    try:
        selected_ref_bytes = [image_bytes_list[valid_indices[int(idx) - 1]] for idx in selected_indexes]
    except Exception as exc:
        raise PipelineStepError(f"reference image index mapping failed: {exc}", {
            "selected_indexes": selected_indexes,
            "valid_indices": valid_indices,
        }) from exc

    log(f"reference images: {len(selected_ref_bytes)} (raw indexes: {[valid_indices[idx - 1] + 1 for idx in selected_indexes]})")
    edit_image = build_edit_image(selected_ref_bytes)

    log(f">>> parallel generating {len(prompt_items)} images (max {MAX_PARALLEL} concurrent)...")
    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_results = []
    _ok_lock = threading.Lock()
    _stats = {"ok": 0, "fail": 0}

    worker_count = min(MAX_PARALLEL, len(prompt_items))

    def _process_one(task_index: int, task_total: int, task_number: int,
                     image_prompt: str) -> dict:
        task_name = f"image_{task_number}"
        log(f"[{task_index}/{task_total}] {task_name} start")
        log(f"  prompt: {image_prompt[:120]}...")

        try:
            image_url, oss_result, elapsed, attempts = generate_one_image(
                env=env,
                task_name=task_name,
                prompt=image_prompt,
                api_key=vibe_api_key,
                base_url=vibe_base_url,
                edit_image=edit_image,
                size=size,
                model=image_model,
            )
            with _ok_lock:
                _stats["ok"] += 1
            log(f"[{task_index}/{task_total}] {task_name} OK ({elapsed:.1f}s, {attempts} attempts)")
            return {
                "source": f"generated_{task_number}",
                "image_type": task_name,
                "generated_image": image_url,
                "oss_object_key": oss_result.get("object_key", ""),
                "prompt": image_prompt,
                "error": None,
                "elapsed": elapsed,
            }
        except Exception as exc:
            with _ok_lock:
                _stats["fail"] += 1
            log(f"[{task_index}/{task_total}] {task_name} FAILED: {exc}")
            log(f"  Traceback: {_tb.format_exc()}")
            return {
                "source": f"generated_{task_number}",
                "image_type": task_name,
                "generated_image": None,
                "prompt": image_prompt,
                "error": str(exc),
                "elapsed": 0,
            }

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = []
        for idx, (task_number, image_prompt) in enumerate(prompt_items, start=1):
            futures.append(
                executor.submit(_process_one, idx, len(prompt_items),
                                task_number, image_prompt)
            )

        for future in as_completed(futures):
            generated_results.append(future.result())

    # sort by task_number
    generated_results.sort(key=lambda r: int(r["image_type"].split("_")[1])
                           if "_" in r.get("image_type", "") else 0)

    generation_stats = {
        "total_input_images": image_context["total_input_images"],
        "valid_images": image_context["valid_images"],
        "old_image_count": len(old_image_urls),
        "new_image_count": sum(1 for item in generated_results if item.get("generated_image")),
        "selected_indexes": selected_indexes,
        "total_generated": len(prompt_items),
        "success": _stats["ok"],
        "failed": _stats["fail"],
        "model": image_model,
        "size": size,
    }

    log(f">>> STEP4 done: success {_stats['ok']} images, failed {_stats['fail']} images")

    # save metadata
    meta_path = OUTPUT_DIR / f"step4_generation_{run_ts}.json"
    meta_path.write_text(json.dumps({
        "en_title": vision_result.get("en_title", ""),
        "selected_indexes": selected_indexes,
        "old_image": old_image_urls,
        "new_image": [item["generated_image"] for item in generated_results if item.get("generated_image")],
        "generation_stats": generation_stats,
        "results": generated_results,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "step": "step4_generation",
        "generated": generated_results,
        "generation_stats": generation_stats,
        "old_image": old_image_urls,
        "new_image": [item["generated_image"] for item in generated_results if item.get("generated_image")],
        "meta_path": str(meta_path),
    }


def step3_generate_images(
    env: dict[str, str],
    products: list[dict],
    progress_callback=None,
) -> dict:
    """Compat wrapper: run Step3 Vision + Step4 generation sequentially."""
    vision_result = step3_analyze_vision(env, products, progress_callback=progress_callback)
    generation_result = step4_generate_images(
        env,
        products,
        vision_result,
        progress_callback=progress_callback,
    )
    return {
        "generated": generation_result.get("generated", []),
        "vision_stats": generation_result.get("generation_stats", {}),
        "vision": vision_result,
        "generation": generation_result,
    }


def export_to_xlsx(raw_import: dict, products: list[dict], output_path: str) -> str:
    """Export the full 60-column xlsx."""
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "popTemu_product"

    headers = [
        '产品标题','英文标题','产品描述','产品货号','变种名称',
        '变种属性名称一','变种属性值一','变种属性名称二','变种属性值二','预览图',
        '申报价格','SKU货号','长','宽','高',
        '重量','识别码类型','识别码','站外产品链接','轮播图',
        '产品素材图','外包装形状','外包装类型','外包装图片','建议零售价','建议零售价币种',
        '库存','发货时效','分类id','产品属性','SPU属性',
        'SKC属性','SKU属性','站点价格','来源url','产地',
        '敏感属性','备注','SKU分类','SKU分类数量','SKU分类单位',
        '独立包装','净含量数值','净含量单位','混合套装类型','SKU分类总数量',
        'SKU分类总数量单位','总净含量','总净含量单位','包装清单','生命周期',
        '视频Url','运费模板（模板id）','经营站点','所属店铺','SPUID',
        'SKCID','SKUID','创建时间','更新时间'
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    shop = raw_import.get("shopConfig", {})
    goods_id = raw_import.get("goodsId", "")
    category_id = raw_import.get("categoryId", "")
    video_url = raw_import.get("videoUrl", "")
    created_at = raw_import.get("createdAt", "")
    skus = raw_import.get("skus", [])
    product_data = raw_import.get("product", {})

    gallery_images = product_data.get("galleryImages", [])
    gallery_str = "\n".join(gallery_images)
    first_img = gallery_images[0] if gallery_images else ""

    raw_props = product_data.get("productProps", [])
    clean_props = []
    for p in raw_props:
        clean_props.append({
            "propName": p.get("propName", ""),
            "refPid": p.get("refPid", ""),
            "pid": p.get("pid", ""),
            "templatePid": p.get("templatePid", ""),
            "numberInputValue": p.get("numberInputValue", ""),
            "valueUnit": p.get("valueUnit", ""),
            "vid": p.get("vid", ""),
            "propValue": p.get("propValue", ""),
        })
    props_json = json.dumps(clean_props, ensure_ascii=False)

    cn_title = product_data.get("title", "")
    en_title = ""
    if products:
        cn_title = products[0].get("chinese_title", product_data.get("title", ""))
        en_title = products[0].get("english_title", "")

    generated_urls = []
    if products:
        gen_list = products[0].get("generated", [])
        for g in gen_list:
            gen_path = g.get("generated_image", "")
            if gen_path:
                generated_urls.append(gen_path)

    def fmt_decimal(v):
        if v is None or v == "":
            return ""
        try:
            return f"{float(v):.1f}"
        except (ValueError, TypeError):
            return ""

    def clean_price(p):
        if not p:
            return ""
        s = str(p).replace("$", "").replace("¥", "").replace(",", "").replace(" ", "")
        try:
            return str(float(s))
        except ValueError:
            return ""

    row_idx = 2
    for sku in skus:
        first_generated = generated_urls[0] if generated_urls else ""
        gen_str = "\n".join(generated_urls)

        r = [""] * 59
        r[0]  = cn_title
        r[1]  = en_title
        r[4]  = sku.get("variantName", "")
        r[5]  = sku.get("specName1", "")
        r[6]  = sku.get("specValue1", "")
        r[7]  = sku.get("specName2", "")
        r[8]  = sku.get("specValue2", "")
        r[9]  = first_generated
        r[10] = clean_price(shop.get("declarePrice") or sku.get("price", ""))
        r[12] = fmt_decimal(shop.get("length", ""))
        r[13] = fmt_decimal(shop.get("width", ""))
        r[14] = fmt_decimal(shop.get("height", ""))
        r[15] = fmt_decimal(shop.get("weight", ""))
        r[19] = gen_str
        r[20] = first_generated
        r[24] = clean_price(shop.get("retailPrice") or sku.get("price", ""))
        r[25] = shop.get("stock") or sku.get("stock", 0)
        r[26] = "9"
        r[27] = category_id
        r[28] = props_json
        r[29] = "[]"
        r[30] = sku.get("skcProps", "[]")
        r[31] = sku.get("skuProps", "[]")
        r[34] = shop.get("origin", "")
        r[37] = shop.get("skuClass") or "按件包装"
        r[38] = shop.get("skuClassQty") or "7"
        r[39] = shop.get("skuClassUnit") or "件"
        r[41] = "0"
        r[44] = "0"
        r[46] = "0"
        r[50] = video_url
        r[51] = shop.get("shipping", "")
        r[52] = shop.get("site", "")
        r[53] = shop.get("shopName", "")
        r[54] = sku.get("spuId", goods_id)
        r[55] = sku.get("skcId", "")
        r[56] = sku.get("skuId", "")
        r[57] = created_at
        r[58] = created_at

        for col, val in enumerate(r, 1):
            ws.cell(row=row_idx, column=col, value=val)
        row_idx += 1

    col_widths = [
        40, 40, 30, 15, 20, 15, 20, 15, 20, 50,
        12, 18, 8, 8, 8, 8, 12, 15, 50, 50,
        50, 12, 12, 50, 15, 8, 10, 10, 60, 10,
        50, 50, 12, 50, 20, 12, 30, 12, 10, 10,
        10, 10, 10, 10, 10, 10, 10, 10, 40, 12,
        50, 20, 20, 15, 20, 20, 20, 20, 20
    ]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    return output_path
