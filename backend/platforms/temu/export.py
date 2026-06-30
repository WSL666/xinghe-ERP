"""Temu 导出:生成符合 popTemu 模板的 59 列 xlsx。

支持单条导出(to_xlsx)和批量导出(to_xlsx_batch)。
批量导出时所有链接的 SKU 行写入同一个 sheet,数据连续不断,不插空行。
"""
from __future__ import annotations

import json
from typing import Any


HEADERS = [
    '产品标题','英文标题','产品描述','产品货号','变种名称',
    '变种属性名称一','变种属性值一','变种属性名称二','变种属性值二','预览图',
    '申报价格','SKU货号','长','宽','高',
    '重量','识别码类型','识别码','站外产品链接','轮播图',
    '产品素材图','外包装形状','外包装类型','外包装图片','建议零售价(建议零售价币种)',
    '库存','发货时效','分类id','产品属性','SPU属性',
    'SKC属性','SKU属性','站点价格','来源url','产地',
    '敏感属性','备注','SKU分类','SKU分类数量','SKU分类单位',
    '独立包装','净含量数值','净含量单位','混合套装类型','SKU分类总数量',
    'SKU分类总数量单位','总净含量','总净含量单位','包装清单','生命周期',
    '视频Url','运费模板（模板id）','经营站点','所属店铺','SPUID',
    'SKCID','SKUID','创建时间','更新时间'
]

COL_WIDTHS = [
    40, 40, 30, 15, 20, 15, 20, 15, 20, 50,
    12, 18, 8, 8, 8, 8, 12, 15, 50, 50,
    50, 12, 12, 50, 15, 8, 10, 10, 60, 10,
    50, 50, 12, 50, 20, 12, 30, 12, 10, 10,
    10, 10, 10, 10, 10, 10, 10, 10, 40, 12,
    50, 20, 20, 15, 20, 20, 20, 20, 20
]


def _fmt_decimal(v):
    if v is None or v == "":
        return ""
    try:
        return f"{float(v):.1f}"
    except (ValueError, TypeError):
        return ""


def _clean_price(p):
    if not p:
        return ""
    s = str(p).replace("$", "").replace("¥", "").replace(",", "").replace(" ", "")
    try:
        return str(float(s))
    except ValueError:
        return ""


def _build_sku_rows(raw_import: dict[str, Any], cn_title: str, en_title: str,
                    generated: list[dict[str, Any]]) -> list[list]:
    """根据一条采集数据构建其所有 SKU 行(每行 59 列的 list)。"""
    shop = raw_import.get("shopConfig", {})
    goods_id = raw_import.get("goodsId", "")
    category_id = raw_import.get("categoryId", "")
    video_url = raw_import.get("videoUrl", "")
    created_at = raw_import.get("createdAt", "")
    skus = raw_import.get("skus", [])
    product_data = raw_import.get("product", {})

    raw_props = product_data.get("productProps", [])
    clean_props = [{
        "propName": p.get("propName", ""), "refPid": p.get("refPid", ""),
        "pid": p.get("pid", ""), "templatePid": p.get("templatePid", ""),
        "numberInputValue": p.get("numberInputValue", ""), "valueUnit": p.get("valueUnit", ""),
        "vid": p.get("vid", ""), "propValue": p.get("propValue", ""),
    } for p in raw_props]
    props_json = json.dumps(clean_props, ensure_ascii=False)

    pack_list = product_data.get("packList", []) or []
    pack_list_json = json.dumps(pack_list, ensure_ascii=False) if pack_list else ""

    generated_urls = [g.get("generated_image", "") for g in generated if g.get("generated_image")]
    gen_str = "\n".join(generated_urls)
    first_generated = generated_urls[0] if generated_urls else ""

    rows = []
    for sku in skus:
        r = [""] * 59
        r[0]  = cn_title or product_data.get("title", "")
        r[1]  = en_title
        r[4]  = sku.get("variantName", "")
        r[5]  = sku.get("specName1", "")
        r[6]  = sku.get("specValue1", "")
        r[7]  = sku.get("specName2", "")
        r[8]  = sku.get("specValue2", "")
        r[9]  = first_generated
        r[10] = _clean_price(shop.get("declarePrice"))
        r[12] = _fmt_decimal(shop.get("length", ""))
        r[13] = _fmt_decimal(shop.get("width", ""))
        r[14] = _fmt_decimal(shop.get("height", ""))
        r[15] = _fmt_decimal(shop.get("weight", ""))
        r[19] = gen_str
        r[20] = first_generated
        r[24] = _clean_price(shop.get("retailPrice"))
        r[25] = shop.get("stock", "")
        r[26] = shop.get("deliveryDays", "")
        r[27] = category_id
        r[28] = props_json
        r[29] = "[]"
        r[30] = sku.get("skcProps", "[]")
        r[31] = sku.get("skuProps", "[]")
        r[34] = shop.get("origin", "")
        r[37] = shop.get("skuClass", "")
        r[38] = shop.get("skuClassQty", "")
        r[39] = shop.get("skuClassUnit", "")
        r[40] = "否"
        r[41] = "0"
        r[44] = "0"
        r[46] = "0"
        r[48] = pack_list_json
        r[50] = video_url
        r[51] = shop.get("shippingTemplateId") or shop.get("shipping", "")
        r[52] = shop.get("site", "")
        r[53] = shop.get("shopName", "")
        r[54] = sku.get("spuId", goods_id)
        r[55] = sku.get("skcId", "")
        r[56] = sku.get("skuId", "")
        r[57] = created_at
        r[58] = created_at
        rows.append(r)
    return rows


def _new_workbook():
    """创建带表头的 workbook,返回 (wb, ws)。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "popTemu_product"
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    return wb, ws


def _save_bytes(wb) -> bytes:
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_xlsx(raw_import: dict[str, Any], cn_title: str, en_title: str,
            generated: list[dict[str, Any]]) -> bytes:
    """单条导出:返回 xlsx bytes。"""
    wb, ws = _new_workbook()
    row_idx = 2
    for r in _build_sku_rows(raw_import, cn_title, en_title, generated):
        for col, val in enumerate(r, 1):
            ws.cell(row=row_idx, column=col, value=val)
        row_idx += 1
    return _save_bytes(wb)


def to_xlsx_batch(items: list[dict[str, Any]]) -> bytes:
    """批量导出:多条链接连续写入同一个 sheet,返回 xlsx bytes。

    items: [{"raw_import": ..., "cn_title": ..., "en_title": ..., "generated": ...}, ...]
    每条链接的所有 SKU 行连续排列,不插空行,不同链接紧挨着。
    """
    wb, ws = _new_workbook()
    row_idx = 2
    for it in items:
        rows = _build_sku_rows(
            it.get("raw_import", {}),
            it.get("cn_title", ""),
            it.get("en_title", ""),
            it.get("generated", []),
        )
        for r in rows:
            for col, val in enumerate(r, 1):
                ws.cell(row=row_idx, column=col, value=val)
            row_idx += 1
    return _save_bytes(wb)
