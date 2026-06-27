"""
product_pipeline_v2 core pipeline

Steps:
Step1: receive plugin data (download image URLs locally) / xlsx upload
Step2: DeepSeek title translation (Chinese <-> English)
Step3: Vision one-shot analysis (select reference images + produce prompts) -> I2I parallel generation
Export: 60-column xlsx
"""

from __future__ import annotations

import base64
import hashlib
import importlib.util
import json
import mimetypes
import re
import sys
import threading
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
from openai import APITimeoutError, OpenAI

# path constants
PIPELINE_ROOT = Path(__file__).resolve().parent
APP_ROOT = PIPELINE_ROOT.parent
ENV_PATH = PIPELINE_ROOT / ".env"
OUTPUT_DIR = PIPELINE_ROOT / "output"
PROMPTS_DIR = PIPELINE_ROOT / "prompts"
TEMP_DIR = PIPELINE_ROOT / "temp"

# image-generation API constants
VIBE_OUTPUT_FORMAT = "png"
VIBE_RESPONSE_FORMAT = "b64_json"

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp"}
MAX_PARALLEL = 10
IMAGE_DOWNLOAD_CONCURRENCY = 8
IMAGE_ATTEMPT_TIMEOUT = 150.0
MAX_IMAGE_ATTEMPTS = 3
IMAGE_DOWNLOAD_TIMEOUT = 60.0
VISION_TIMEOUT = 300.0
VISION_MAX_ATTEMPTS = 3

_print_lock = threading.Lock()


def log(message: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    safe = message.encode("ascii", errors="replace").decode("ascii")
    with _print_lock:
        print(f"[{ts}] {safe}", flush=True)


class PipelineStepError(RuntimeError):
    """Step exception carrying structured detail for the upper layer to write to DB logs."""

    def __init__(self, message: str, detail: dict[str, Any] | None = None):
        super().__init__(message)
        self.detail = detail or {}


def load_env(path: Path = ENV_PATH) -> dict[str, str]:
    if not path.exists():
        raise FileNotFoundError(f"config file not found: {path}")
    env: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip()
    return env


def require_env(env: dict[str, str], key: str) -> str:
    value = env.get(key, "").strip()
    if not value:
        raise ValueError(f"missing required .env config: {key}")
    return value


def parse_json_response(text: str) -> Any:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("{")
        end = text.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise
        return json.loads(text[start : end + 1])


def load_prompt_module(prompt_file: str) -> str:
    path = PROMPTS_DIR / prompt_file
    if not path.exists():
        raise FileNotFoundError(f"prompt file not found: {path}")
    spec = importlib.util.spec_from_file_location(prompt_file.rstrip(".py"), path)
    if spec is None or spec.loader is None:
        raise ImportError(f"cannot load module: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    prompt = getattr(module, "PROMPT", None)
    if not isinstance(prompt, str) or not prompt.strip():
        raise ValueError(f"define a non-empty PROMPT string in {path}")
    return prompt.strip()


def download_bytes(url: str) -> tuple[bytes, str]:
    """Download an image, return (bytes, mime_type)."""
    req = urllib.request.Request(url, method="GET", headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    })
    with urllib.request.urlopen(req, timeout=60) as resp:
        img_bytes = resp.read()
    mime = resp.headers.get("Content-Type", "image/jpeg")
    if ";" in mime:
        mime = mime.split(";")[0]
    return img_bytes, mime


def bytes_to_data_uri(img_bytes: bytes, mime: str) -> str:
    return f"data:{mime};base64,{base64.b64encode(img_bytes).decode('ascii')}"


def guess_mime_bytes(img_bytes: bytes) -> str:
    """Sniff MIME type from file header."""
    if img_bytes[:4] == b'\x89PNG':
        return "image/png"
    if img_bytes[:2] == b'\xff\xd8':
        return "image/jpeg"
    if img_bytes[:4] == b'RIFF' and img_bytes[8:12] == b'WEBP':
        return "image/webp"
    return "image/jpeg"


def encode_image_data_url(img_bytes: bytes) -> str:
    mime = guess_mime_bytes(img_bytes)
    encoded = base64.standard_b64encode(img_bytes).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def image_suffix_for_mime(mime: str) -> str:
    mime = (mime or "").split(";", 1)[0].lower().strip()
    return {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/webp": ".webp",
        "image/gif": ".gif",
    }.get(mime, ".jpg")


def upload_image_bytes_to_oss(
    env: dict[str, str],
    img_bytes: bytes,
    filename: str,
    mime: str,
    folder: str,
) -> dict[str, Any]:
    from oss_client import upload_image_bytes

    result = upload_image_bytes(
        env,
        img_bytes,
        filename=filename,
        content_type=mime,
        folder=folder,
    )
    return {
        "ok": result.ok,
        "url": result.url,
        "object_key": result.object_key,
        "bucket": result.bucket,
        "folder": result.folder,
        "size": result.size,
        "content_type": result.content_type,
    }


def upload_old_image_to_oss(env: dict[str, str], img_bytes: bytes, mime: str, index: int) -> dict[str, Any]:
    suffix = image_suffix_for_mime(mime or guess_mime_bytes(img_bytes))
    filename = f"old_image_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{index}{suffix}"
    return upload_image_bytes_to_oss(env, img_bytes, filename, mime or guess_mime_bytes(img_bytes), folder="1")


def upload_new_image_to_oss(env: dict[str, str], img_bytes: bytes, task_name: str) -> dict[str, Any]:
    jpg_bytes = ensure_jpeg_bytes(img_bytes)
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{task_name}.jpg"
    return upload_image_bytes_to_oss(env, jpg_bytes, filename, "image/jpeg", folder="2")


def ensure_jpeg_bytes(img_bytes: bytes) -> bytes:
    """Return JPEG bytes for generated images before OSS upload."""
    try:
        from PIL import Image
        from io import BytesIO

        source = BytesIO(img_bytes)
        with Image.open(source) as image:
            if image.mode in {"RGBA", "LA"}:
                background = Image.new("RGB", image.size, (255, 255, 255))
                alpha = image.getchannel("A") if "A" in image.getbands() else None
                background.paste(image.convert("RGBA"), mask=alpha)
                output_image = background
            else:
                output_image = image.convert("RGB")

            output = BytesIO()
            output_image.save(output, format="JPEG", quality=95, optimize=True)
            return output.getvalue()
    except ImportError as exc:
        raise RuntimeError("Pillow is required to convert generated images to JPG") from exc


def upload_source_image_urls_to_oss(env: dict[str, str], urls: list[str]) -> list[str]:
    old_image_urls: list[str] = []
    for index, url in enumerate(urls[:10], start=1):
        raw_bytes, mime = download_bytes(url)
        oss_result = upload_old_image_to_oss(env, raw_bytes, mime, index)
        old_image_urls.append(oss_result["url"])
    return old_image_urls


def video_suffix_for_mime(mime: str) -> str:
    mime = (mime or "").split(";", 1)[0].lower().strip()
    return {
        "video/mp4": ".mp4",
        "video/webm": ".webm",
        "video/quicktime": ".mov",
        "video/x-msvideo": ".avi",
        "video/x-matroska": ".mkv",
    }.get(mime, ".mp4")


def upload_file_bytes_to_oss(
    env: dict[str, str],
    file_bytes: bytes,
    filename: str,
    mime: str,
    folder: str,
) -> dict[str, Any]:
    from oss_client import upload_file_bytes

    result = upload_file_bytes(
        env,
        file_bytes,
        filename=filename,
        content_type=mime,
        folder=folder,
    )
    return {
        "ok": result.ok,
        "url": result.url,
        "object_key": result.object_key,
        "bucket": result.bucket,
        "folder": result.folder,
        "size": result.size,
        "content_type": result.content_type,
    }


def upload_old_video_to_oss(env: dict[str, str], video_bytes: bytes, mime: str, index: int) -> dict[str, Any]:
    suffix = video_suffix_for_mime(mime)
    filename = f"old_video_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{index}{suffix}"
    return upload_file_bytes_to_oss(env, video_bytes, filename, mime or "video/mp4", folder="3")


def upload_source_videos_to_oss(env: dict[str, str], videos: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Download each source video and upload to OSS old_video folder.

    Videos are display-only (independent of AI image generation). Failures for a
    single video are logged and skipped so a bad URL never breaks the pipeline.
    Returns the videos with an added `oss_url` field.
    """
    result: list[dict[str, Any]] = []
    for index, video in enumerate(videos, start=1):
        if isinstance(video, str):
            video = {"url": video, "poster": "", "width": 0, "height": 0}
        entry = {**video}
        url = video.get("url") or ""
        if not url:
            result.append(entry)
            continue
        try:
            raw_bytes, mime = download_bytes(url)
            oss_result = upload_old_video_to_oss(env, raw_bytes, mime, index)
            entry["oss_url"] = oss_result["url"]
        except Exception as exc:  # noqa: BLE001 - best effort, display-only
            log(f"[WARN] video {index} upload failed, kept original url {url}: {str(exc)[:160]}")
            entry["oss_url"] = ""
        result.append(entry)
    return result


def call_text_llm(env: dict[str, str], prompt_str: str, max_tokens: int = 4096,
                  base_url: str = None, api_key: str = None, model: str = None) -> str:
    """Call a text-only OpenAI-compatible LLM (e.g. DeepSeek) via the SDK."""
    _api_key = api_key or require_env(env, "step2_api_key")
    _base_url = (base_url or require_env(env, "step2_base_url")).rstrip("/")
    if _base_url.endswith("/chat/completions"):
        _base_url = _base_url[: -len("/chat/completions")]
    _model = model or env.get("step2_model", "deepseek-chat")

    log(f"text LLM: model={_model}, base={_base_url}")
    client = OpenAI(base_url=_base_url, api_key=_api_key)
    resp = client.chat.completions.create(
        model=_model,
        messages=[{"role": "user", "content": prompt_str}],
        max_tokens=max_tokens,
        timeout=120,
    )
    content = resp.choices[0].message.content
    if not isinstance(content, str) or not content.strip():
        raise RuntimeError("LLM API returned empty content")
    return content.strip()


def image_id(source: str) -> str:
    return hashlib.md5(source.encode()).hexdigest()[:12]


def natural_key(path: Path) -> list[object]:
    return [int(part) if part.isdigit() else part.lower()
            for part in re.split(r"(\d+)", path.name)]


def build_vision_messages(prompt: str, image_b64_list: list[str]) -> list[dict[str, Any]]:
    """Build OpenAI SDK multimodal messages."""
    content: list[dict[str, Any]] = [{"type": "text", "text": prompt}]
    for b64_url in image_b64_list:
        content.append({
            "type": "image_url",
            "image_url": {"url": b64_url},
        })
    return [{"role": "user", "content": content}]


def analyze_product(env: dict[str, str], prompt: str,
                    image_b64_list: list[str]) -> dict[str, Any]:
    """Call the Vision model to analyze all images in one shot -> output JSON."""
    import traceback as _tb
    chat_api_key = require_env(env, "CHAT_API_KEY")
    chat_base_url = require_env(env, "OPENAI_CHAT_BASE_URL")
    chat_model = env.get("CHAT_MODEL", "gpt-5.5")

    if chat_base_url.endswith("/chat/completions"):
        sdk_base = chat_base_url[:-len("/chat/completions")]
    else:
        sdk_base = chat_base_url.rstrip("/")

    log(f"Vision: base={sdk_base}, model={chat_model}, key=...{chat_api_key[-8:]}, images={len(image_b64_list)}")

    client = OpenAI(
        base_url=sdk_base,
        api_key=chat_api_key,
        default_headers={"User-Agent": "python-httpx/0.28.1"},
    )
    try:
        stream = client.chat.completions.create(
            model=chat_model,
            messages=build_vision_messages(prompt, image_b64_list),
            stream=True,
            stream_options={"include_usage": True},
            timeout=httpx.Timeout(VISION_TIMEOUT, connect=30.0),
        )
        content_parts: list[str] = []
        usage = None
        for chunk in stream:
            if getattr(chunk, "usage", None):
                usage = chunk.usage
            choices = getattr(chunk, "choices", None) or []
            if not choices:
                continue
            delta = getattr(choices[0], "delta", None)
            delta_content = getattr(delta, "content", None) if delta is not None else None
            if delta_content:
                content_parts.append(delta_content)
    except Exception as exc:
        log(f"Vision API call exception: {exc}")
        log(f"Traceback: {_tb.format_exc()}")
        raise RuntimeError(f"Vision API call failed: {exc}") from exc

    content = "".join(content_parts).strip()
    if not content:
        raise RuntimeError("Empty vision stream response")
    if usage:
        log(f"Vision usage: {usage}")
    log(f"Vision raw response (first 500 chars): {content[:500]}")
    return parse_json_response(content)


def validate_analysis_payload(payload: dict[str, Any],
                               image_count: int) -> tuple[list[int], list[tuple[int, str]]]:
    """Validate Vision result."""
    raw_indexes = payload.get("selected_reference_image_indexes")
    if not isinstance(raw_indexes, list):
        raise ValueError("selected_reference_image_indexes must be a list")

    indexes: list[int] = []
    for item in raw_indexes:
        if isinstance(item, bool):
            raise ValueError("selected_reference_image_indexes cannot contain booleans")
        try:
            index = int(item)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid reference image index: {item!r}") from exc
        indexes.append(index)

    if len(indexes) < 2 or len(indexes) > 3:
        raise ValueError("selected_reference_image_indexes must contain 2 to 3 indexes")
    if len(set(indexes)) != len(indexes):
        raise ValueError("selected_reference_image_indexes contains duplicate indexes")
    invalid = [index for index in indexes if index < 1 or index > image_count]
    if invalid:
        raise ValueError(f"Reference image indexes out of range 1..{image_count}: {invalid}")

    prompt_items: list[tuple[int, str]] = []
    for key, value in payload.items():
        match = re.fullmatch(r"image_(\d+)", key)
        if not match:
            continue
        if not isinstance(value, str) or not value.strip():
            raise ValueError(f"{key} must be a non-empty string prompt")
        prompt_items.append((int(match.group(1)), value.strip()))

    prompt_items.sort(key=lambda item: item[0])
    prompt_numbers = [number for number, _ in prompt_items]
    if len(prompt_items) < 6 or len(prompt_items) > 8:
        raise ValueError(f"Expected 6 to 8 image_N prompts, got {len(prompt_items)}")
    if prompt_numbers != list(range(1, len(prompt_items) + 1)):
        raise ValueError(f"image_N keys must be continuous from image_1: {prompt_numbers}")

    return indexes, prompt_items


def _download_one_image(index: int, url: str, total: int) -> tuple[int, dict[str, Any]]:
    """Download a single carousel image. Returns (index, item) so results stay
    index-aligned regardless of completion order."""
    item = {"index": index + 1, "url": url, "ok": False, "bytes": 0, "mime": "", "error": ""}
    try:
        log(f"  download [{index + 1}/{total}]: {url[:80]}...")
        raw_bytes, mime = download_bytes(url)
        item.update({"ok": True, "bytes": len(raw_bytes), "mime": mime, "_raw": raw_bytes})
        log(f"    ok [{index + 1}]: {len(raw_bytes)} bytes, mime={mime}")
    except Exception as exc:
        item["error"] = str(exc)
        log(f"    fail [{index + 1}]: {exc}")
    return index, item


def _download_images_parallel(all_urls: list[str]) -> tuple[list, list, list, list[int]]:
    """Download all carousel images in parallel, preserving input order.

    Replaces the serial download loop that was the dominant wall-clock cost in
    step3 (up to 10 sequential round-trips). A bounded thread pool keeps memory
    in check; per-image failures stay isolated exactly as before (None entries
    keep the lists index-aligned with all_urls).
    """
    total = len(all_urls)
    workers = min(IMAGE_DOWNLOAD_CONCURRENCY, total)
    fetched: dict[int, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(_download_one_image, i, url, total) for i, url in enumerate(all_urls)]
        for fut in as_completed(futures):
            idx, item = fut.result()
            fetched[idx] = item
    image_bytes_list: list[bytes | None] = [None] * total
    image_b64_list: list[str | None] = [None] * total
    downloads: list[dict[str, Any]] = []
    for i in range(total):
        item = fetched[i]
        raw = item.pop("_raw", None)
        if item["ok"] and raw is not None:
            image_bytes_list[i] = raw
            image_b64_list[i] = encode_image_data_url(raw)
        downloads.append(item)
    valid_indices = [i for i, b in enumerate(image_bytes_list) if b is not None]
    valid_b64 = [b for b in image_b64_list if b is not None]
    return image_bytes_list, image_b64_list, downloads, valid_indices


def collect_product_images(products: list[dict]) -> dict[str, Any]:
    """Download and encode product carousel images for Vision and I2I reuse."""
    all_urls = []
    for product in products:
        for img_url in product.get("carousel_images", [])[:10]:
            all_urls.append(img_url)

    if not all_urls:
        raise PipelineStepError("no carousel images to process", {
            "total_input_images": 0,
            "valid_images": 0,
            "downloads": [],
        })

    log(f"total {len(all_urls)} carousel image URLs to download")
    image_bytes_list, image_b64_list, downloads, valid_indices = _download_images_parallel(all_urls)
    valid_b64 = [b for b in image_b64_list if b is not None]
    log(f"download complete: {len(valid_b64)}/{len(all_urls)} usable")

    result = {
        "all_urls": all_urls,
        "image_bytes_list": image_bytes_list,
        "image_b64_list": image_b64_list,
        "valid_b64": valid_b64,
        "valid_indices": valid_indices,
        "downloads": downloads,
        "total_input_images": len(all_urls),
        "valid_images": len(valid_b64),
    }

    if not valid_b64:
        raise PipelineStepError("all images download failed", result)

    return result


def summarize_image_inputs(image_context: dict[str, Any]) -> dict[str, Any]:
    return {
        "total_input_images": image_context.get("total_input_images", 0),
        "valid_images": image_context.get("valid_images", 0),
        "downloads": image_context.get("downloads", []),
    }
def analyze_product_with_retry(
    env: dict[str, str],
    vision_prompt: str,
    valid_b64: list[str],
    image_count: int,
    max_attempts: int = VISION_MAX_ATTEMPTS,
) -> dict[str, Any]:
    """Vision analysis with retry; validates JSON schema each attempt."""
    attempts: list[dict[str, Any]] = []
    last_error = ""

    for attempt in range(1, max_attempts + 1):
        started = time.perf_counter()
        attempt_info: dict[str, Any] = {
            "attempt": attempt,
            "ok": False,
            "elapsed": 0,
            "error": "",
            "payload_preview": "",
        }
        try:
            payload = analyze_product(env, vision_prompt, valid_b64)
            attempt_info["payload_preview"] = json.dumps(payload, ensure_ascii=False)[:1000]
            selected_indexes, prompt_items = validate_analysis_payload(payload, image_count)
            attempt_info.update({
                "ok": True,
                "elapsed": round(time.perf_counter() - started, 3),
                "selected_indexes": selected_indexes,
                "prompt_count": len(prompt_items),
            })
            attempts.append(attempt_info)
            return {
                "payload": payload,
                "selected_indexes": selected_indexes,
                "prompt_items": prompt_items,
                "attempts": attempts,
            }
        except Exception as exc:
            last_error = str(exc)
            attempt_info.update({
                "elapsed": round(time.perf_counter() - started, 3),
                "error": last_error,
            })
            attempts.append(attempt_info)
            log(f"[WARN] Vision attempt {attempt}/{max_attempts} failed: {exc}")
            if attempt < max_attempts:
                time.sleep(min(2 * attempt, 6))

    raise PipelineStepError(f"Vision analysis failed after {max_attempts} retries: {last_error}", {
        "attempts": attempts,
        "last_error": last_error,
    })


def build_edit_image(image_bytes_list: list[bytes]):
    """Build the image argument for OpenAI SDK images.edit."""
    files = []
    for i, img_bytes in enumerate(image_bytes_list):
        mime = guess_mime_bytes(img_bytes)
        fname = f"ref_{i + 1}.{mime.split('/')[-1]}"
        files.append((fname, img_bytes, mime))
    if len(files) == 1:
        return files[0]
    return files


def create_vibe_client(api_key: str, base_url: str) -> OpenAI:
    return OpenAI(
        base_url=base_url,
        api_key=api_key,
        default_headers={"User-Agent": "python-httpx/0.28.1"},
    )


def read_result_item_bytes(item: Any, timeout: float) -> bytes:
    if item.b64_json:
        return base64.b64decode(item.b64_json)

    if item.url:
        response = httpx.get(item.url, timeout=timeout)
        if response.is_error:
            raise RuntimeError(f"HTTP {response.status_code}: {response.text[:500]}")
        return response.content

    raise RuntimeError(f"Could not parse image result: {item}")


def is_timeout_error(exc: Exception) -> bool:
    if isinstance(exc, (APITimeoutError, httpx.TimeoutException)):
        return True
    name = type(exc).__name__.lower()
    return "timeout" in name or "timed out" in str(exc).lower()


def generate_one_image(
    env: dict[str, str],
    task_name: str,
    prompt: str,
    api_key: str,
    base_url: str,
    edit_image: Any,
    size: str,
    model: str,
    attempt_timeout: float = IMAGE_ATTEMPT_TIMEOUT,
    max_attempts: int = MAX_IMAGE_ATTEMPTS,
) -> tuple[str, dict[str, Any], float, int]:
    """Call images.edit to generate one image."""
    started = time.perf_counter()
    last_error = "unknown error"

    for attempt in range(1, max_attempts + 1):
        log(f"{task_name}: attempt {attempt}/{max_attempts}")
        client = create_vibe_client(api_key, base_url)
        try:
            response = client.images.edit(
                image=edit_image,
                prompt=prompt,
                model=model,
                size=size,
                n=1,
                output_format=VIBE_OUTPUT_FORMAT,
                response_format=VIBE_RESPONSE_FORMAT,
                timeout=attempt_timeout,
            )
        except Exception as exc:
            last_error = str(exc)
            if is_timeout_error(exc) and attempt < max_attempts:
                log(f"[WARN] {task_name}: timeout, retrying...")
                continue
            raise

        data = response.data or []
        if not data:
            last_error = "image response has no data"
            if attempt < max_attempts:
                log(f"[WARN] {task_name}: {last_error}, retrying...")
                continue
            raise RuntimeError(f"{task_name}: {last_error}")

        image_bytes = read_result_item_bytes(data[0], IMAGE_DOWNLOAD_TIMEOUT)
        oss_result = upload_new_image_to_oss(env, image_bytes, task_name)
        elapsed = time.perf_counter() - started
        log(f"[OK] {task_name}: uploaded OSS ({elapsed:.2f}s, {attempt} attempt(s))")
        return oss_result["url"], oss_result, elapsed, attempt

    raise RuntimeError(
        f"{task_name} failed after {max_attempts} attempts: {last_error}"
    )
    raise RuntimeError(
        f"{task_name} failed after {max_attempts} attempts: {last_error}"
    )
# Step1: read xlsx

def step1_read_xlsx(xlsx_path: str) -> list[dict]:
    """Read xlsx, extract product title and carousel images."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col

    title_col = None
    carousel_col = None
    for key, col in headers.items():
        if "产品标题" in key or key == "产品标题":
            title_col = col
        if "轮播图" in key or key == "轮播图":
            carousel_col = col

    if title_col is None:
        title_col = 1
    if carousel_col is None:
        carousel_col = 20

    results = []
    for row_idx in range(2, ws.max_row + 1):
        title = ws.cell(row=row_idx, column=title_col).value
        carousel_raw = ws.cell(row=row_idx, column=carousel_col).value
        if not title or not str(title).strip():
            continue

        carousel_images = []
        if carousel_raw:
            carousel_text = str(carousel_raw).strip()
            urls = re.split(r'[\n\r]+', carousel_text)
            carousel_images = [u.strip() for u in urls if u.strip().startswith("http")]
            if len(carousel_images) > 10:
                carousel_images = carousel_images[:10]

        results.append({
            "row": row_idx,
            "chinese_title": str(title).strip(),
            "carousel_images": carousel_images,
        })

    wb.close()
    return results


# Step2: DeepSeek title translation

def step2_translate_titles(env: dict[str, str], products: list[dict]) -> list[dict]:
    """DeepSeek batch title translation."""
    log("=" * 50)
    log(">>> STEP2: DeepSeek title translation start")
    prompt_template = load_prompt_module("step2_translate.py")
    titles = [p["chinese_title"] for p in products]
    titles_json = json.dumps(titles, ensure_ascii=False)

    full_prompt = f"{prompt_template}\n\nInput:\n{titles_json}"

    step2_base_url = env.get("step2_base_url", "").strip() or None
    step2_api_key = env.get("step2_api_key", "").strip() or None
    step2_model = env.get("step2_model", "").strip() or None

    log(f"translate model: {step2_model}  endpoint: {step2_base_url}")
    log(f"original title: {titles[0] if titles else 'N/A'}")

    raw_text = call_text_llm(
        env, full_prompt, max_tokens=4096,
        base_url=step2_base_url, api_key=step2_api_key, model=step2_model,
    )
    log(f"translate raw response (first 300 chars): {raw_text[:300]}")
    translated = parse_json_response(raw_text)

    if not isinstance(translated, list):
        raise ValueError(f"translation result is not an array: {raw_text[:300]}")

    results = []
    for i, item in enumerate(translated):
        results.append({
            "row": products[i]["row"] if i < len(products) else i + 2,
            "chinese_title": item.get("cn_title") or item.get("chinese_title",
                                    products[i]["chinese_title"] if i < len(products) else ""),
            "english_title": item.get("en_title") or item.get("english_title", ""),
            "carousel_images": products[i]["carousel_images"] if i < len(products) else [],
        })
    log(f"translate done: cn={results[0]['chinese_title'][:80]}...")
    log(f"translate done: en={results[0]['english_title'][:80]}...")
    log("<<< STEP2 done")
    return results
    log("<<< STEP2 done")
    return results
# Step3: Vision analysis

def step3_analyze_vision(
    env: dict[str, str],
    products: list[dict],
    progress_callback=None,
) -> dict:
    """Step3: download images, call Vision, validate output JSON."""
    from prompts.vision_prompt import build_prompt as build_vision_prompt

    log("=" * 50)
    log(">>> STEP3: Vision analysis start")

    image_context = collect_product_images(products)
    product = products[0] if products else {}
    old_image_urls = product.get("old_image_urls", [])
    # vision consumes the raw chinese title directly; step2 translation
    # runs in parallel and is intentionally not awaited, so we never
    # block vision waiting for it (saves wall-clock time).
    product_text = product.get("chinese_title", "")
    vision_prompt = build_vision_prompt(product_text)

    log(f"Vision prompt product_text: {product_text[:120]}")
    log(f">>> calling Vision model ({len(image_context['valid_b64'])} images), timeout {VISION_TIMEOUT}s...")
    t0 = time.perf_counter()

    analysis = analyze_product_with_retry(
        env,
        vision_prompt,
        image_context["valid_b64"],
        image_context["valid_images"],
    )
    payload = analysis["payload"]
    selected_indexes = analysis["selected_indexes"]
    prompt_items = analysis["prompt_items"]

    vision_elapsed = time.perf_counter() - t0
    log(f"Vision done ({vision_elapsed:.1f}s): selected_indexes={selected_indexes}, "
        f"prompt_count={len(prompt_items)}")
    for num, pt in prompt_items:
        log(f"  image_{num}: {pt[:100]}...")

    result = {
        "step": "step3_vision",
        "en_title": product_text,
        "vision_prompt": vision_prompt,
        "payload": payload,
        "selected_indexes": selected_indexes,
        "prompt_items": [{"number": n, "prompt": p} for n, p in prompt_items],
        "image_context": {
            **summarize_image_inputs(image_context),
            "old_image_urls": old_image_urls,
        },
        "attempts": analysis["attempts"],
        "elapsed": vision_elapsed,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    meta_path = OUTPUT_DIR / f"step3_vision_{run_ts}.json"
    meta_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    result["meta_path"] = str(meta_path)
    # In-memory image cache for step4 reuse. Not persisted to the meta
    # json (bytes are not JSON-serializable); step4 falls back to a fresh
    # download when the cache is absent (e.g. manual run from the db).
    result["_image_cache"] = {
        "image_bytes_list": image_context["image_bytes_list"],
        "valid_indices": image_context["valid_indices"],
    }

    log("<<< STEP3 Vision done")
    return result
    log("<<< STEP3 Vision done")
    return result
# Step4: I2I generation

def step4_generate_images(
    env: dict[str, str],
    products: list[dict],
    vision_result: dict,
    progress_callback=None,
) -> dict:
    """Step4: run I2I generation based on Step3 Vision result."""
    import traceback as _tb

    log("=" * 50)
    log(">>> STEP4: I2I generation start")

    vibe_api_key = require_env(env, "VIBE_API_KEY")
    vibe_base_url = require_env(env, "VIBE_BASE_URL")
    image_model = env.get("IMAGE_MODEL", "gpt-image-2")
    size = env.get("IMAGE_SIZE", "1024x1024")

    product = products[0] if products else {}
    old_image_urls = product.get("old_image_urls", [])
    selected_indexes = [int(idx) for idx in vision_result.get("selected_indexes", [])]
    raw_prompt_items = vision_result.get("prompt_items", [])

    # Reuse images already downloaded by step3 when available (the auto
    # pipeline hands the vision result over in-memory). Fall back to a
    # fresh download for manual runs where vision_result came from the
    # database without the cache.
    cache = vision_result.get("_image_cache") if isinstance(vision_result, dict) else None
    if cache and cache.get("image_bytes_list") is not None and cache.get("valid_indices") is not None:
        log(f"reuse step3 downloaded images: {len(cache['valid_indices'])}/{len(cache['image_bytes_list'])} usable")
        image_context = {
            "total_input_images": len(cache["image_bytes_list"]),
            "valid_images": len(cache["valid_indices"]),
            "valid_indices": cache["valid_indices"],
            "image_bytes_list": cache["image_bytes_list"],
        }
    else:
        image_context = collect_product_images(products)
        log(f"downloaded images fresh: {len(image_context['valid_indices'])}/{len(image_context['image_bytes_list'])} usable")

    prompt_items: list[tuple[int, str]] = []
    for item in raw_prompt_items:
        if isinstance(item, dict):
            prompt_items.append((int(item["number"]), str(item["prompt"])))
        else:
            prompt_items.append((int(item[0]), str(item[1])))

    if not selected_indexes:
        raise PipelineStepError("Step4 missing Vision selected reference image indexes", {
            "selected_indexes": selected_indexes,
            "prompt_count": len(prompt_items),
        })
    if not prompt_items:
        raise PipelineStepError("Step4 missing Vision generation prompts", {
            "selected_indexes": selected_indexes,
            "prompt_count": 0,
        })

    valid_indices = image_context["valid_indices"]
    image_bytes_list = image_context["image_bytes_list"]
    try:
        selected_ref_bytes = [image_bytes_list[valid_indices[int(idx) - 1]] for idx in selected_indexes]
    except Exception as exc:
        raise PipelineStepError(f"reference image index mapping failed: {exc}", {
            "selected_indexes": selected_indexes,
            "valid_indices": valid_indices,
        }) from exc

    log(f"reference images: {len(selected_ref_bytes)} (raw indexes: {[valid_indices[idx - 1] + 1 for idx in selected_indexes]})")
    edit_image = build_edit_image(selected_ref_bytes)

    log(f">>> parallel generating {len(prompt_items)} images (max {MAX_PARALLEL} concurrent)...")
    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_results = []
    _ok_lock = threading.Lock()
    _stats = {"ok": 0, "fail": 0}

    worker_count = min(MAX_PARALLEL, len(prompt_items))

    def _process_one(task_index: int, task_total: int, task_number: int,
                     image_prompt: str) -> dict:
        task_name = f"image_{task_number}"
        log(f"[{task_index}/{task_total}] {task_name} start")
        log(f"  prompt: {image_prompt[:120]}...")

        try:
            image_url, oss_result, elapsed, attempts = generate_one_image(
                env=env,
                task_name=task_name,
                prompt=image_prompt,
                api_key=vibe_api_key,
                base_url=vibe_base_url,
                edit_image=edit_image,
                size=size,
                model=image_model,
            )
            with _ok_lock:
                _stats["ok"] += 1
            log(f"[{task_index}/{task_total}] {task_name} OK ({elapsed:.1f}s, {attempts} attempts)")
            return {
                "source": f"generated_{task_number}",
                "image_type": task_name,
                "generated_image": image_url,
                "oss_object_key": oss_result.get("object_key", ""),
                "prompt": image_prompt,
                "error": None,
                "elapsed": elapsed,
            }
        except Exception as exc:
            with _ok_lock:
                _stats["fail"] += 1
            log(f"[{task_index}/{task_total}] {task_name} FAILED: {exc}")
            log(f"  Traceback: {_tb.format_exc()}")
            return {
                "source": f"generated_{task_number}",
                "image_type": task_name,
                "generated_image": None,
                "prompt": image_prompt,
                "error": str(exc),
                "elapsed": 0,
            }

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = []
        for idx, (task_number, image_prompt) in enumerate(prompt_items, start=1):
            futures.append(
                executor.submit(_process_one, idx, len(prompt_items),
                                task_number, image_prompt)
            )

        for future in as_completed(futures):
            generated_results.append(future.result())

    # sort by task_number
    generated_results.sort(key=lambda r: int(r["image_type"].split("_")[1])
                           if "_" in r.get("image_type", "") else 0)

    generation_stats = {
        "total_input_images": image_context["total_input_images"],
        "valid_images": image_context["valid_images"],
        "old_image_count": len(old_image_urls),
        "new_image_count": sum(1 for item in generated_results if item.get("generated_image")),
        "selected_indexes": selected_indexes,
        "total_generated": len(prompt_items),
        "success": _stats["ok"],
        "failed": _stats["fail"],
        "model": image_model,
        "size": size,
    }

    log(f">>> STEP4 done: success {_stats['ok']} images, failed {_stats['fail']} images")

    # save metadata
    meta_path = OUTPUT_DIR / f"step4_generation_{run_ts}.json"
    meta_path.write_text(json.dumps({
        "en_title": vision_result.get("en_title", ""),
        "selected_indexes": selected_indexes,
        "old_image": old_image_urls,
        "new_image": [item["generated_image"] for item in generated_results if item.get("generated_image")],
        "generation_stats": generation_stats,
        "results": generated_results,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "step": "step4_generation",
        "generated": generated_results,
        "generation_stats": generation_stats,
        "old_image": old_image_urls,
        "new_image": [item["generated_image"] for item in generated_results if item.get("generated_image")],
        "meta_path": str(meta_path),
    }


def step3_generate_images(
    env: dict[str, str],
    products: list[dict],
    progress_callback=None,
) -> dict:
    """Compat wrapper: run Step3 Vision + Step4 generation sequentially."""
    vision_result = step3_analyze_vision(env, products, progress_callback=progress_callback)
    generation_result = step4_generate_images(
        env,
        products,
        vision_result,
        progress_callback=progress_callback,
    )
    return {
        "generated": generation_result.get("generated", []),
        "vision_stats": generation_result.get("generation_stats", {}),
        "vision": vision_result,
        "generation": generation_result,
    }
# Export xlsx

def export_to_xlsx(raw_import: dict, products: list[dict], output_path: str) -> str:
    """Export the full 60-column xlsx."""
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "popTemu_product"

    headers = [
        '产品标题','英文标题','产品描述','产品货号','变种名称',
        '变种属性名称一','变种属性值一','变种属性名称二','变种属性值二','预览图',
        '申报价格','SKU货号','长','宽','高',
        '重量','识别码类型','识别码','站外产品链接','轮播图',
        '产品素材图','外包装形状','外包装类型','外包装图片','建议零售价','建议零售价币种',
        '库存','发货时效','分类id','产品属性','SPU属性',
        'SKC属性','SKU属性','站点价格','来源url','产地',
        '敏感属性','备注','SKU分类','SKU分类数量','SKU分类单位',
        '独立包装','净含量数值','净含量单位','混合套装类型','SKU分类总数量',
        'SKU分类总数量单位','总净含量','总净含量单位','包装清单','生命周期',
        '视频Url','运费模板（模板id）','经营站点','所属店铺','SPUID',
        'SKCID','SKUID','创建时间','更新时间'
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    shop = raw_import.get("shopConfig", {})
    goods_id = raw_import.get("goodsId", "")
    category_id = raw_import.get("categoryId", "")
    video_url = raw_import.get("videoUrl", "")
    created_at = raw_import.get("createdAt", "")
    skus = raw_import.get("skus", [])
    product_data = raw_import.get("product", {})

    gallery_images = product_data.get("galleryImages", [])
    gallery_str = "\n".join(gallery_images)
    first_img = gallery_images[0] if gallery_images else ""

    raw_props = product_data.get("productProps", [])
    clean_props = []
    for p in raw_props:
        clean_props.append({
            "propName": p.get("propName", ""),
            "refPid": p.get("refPid", ""),
            "pid": p.get("pid", ""),
            "templatePid": p.get("templatePid", ""),
            "numberInputValue": p.get("numberInputValue", ""),
            "valueUnit": p.get("valueUnit", ""),
            "vid": p.get("vid", ""),
            "propValue": p.get("propValue", ""),
        })
    props_json = json.dumps(clean_props, ensure_ascii=False)

    cn_title = product_data.get("title", "")
    en_title = ""
    if products:
        cn_title = products[0].get("chinese_title", product_data.get("title", ""))
        en_title = products[0].get("english_title", "")

    generated_urls = []
    if products:
        gen_list = products[0].get("generated", [])
        for g in gen_list:
            gen_path = g.get("generated_image", "")
            if gen_path:
                generated_urls.append(gen_path)

    def fmt_decimal(v):
        if v is None or v == "":
            return ""
        try:
            return f"{float(v):.1f}"
        except (ValueError, TypeError):
            return ""

    def clean_price(p):
        if not p:
            return ""
        s = str(p).replace("$", "").replace("¥", "").replace(",", "").replace(" ", "")
        try:
            return str(float(s))
        except ValueError:
            return ""

    row_idx = 2
    for sku in skus:
        first_generated = generated_urls[0] if generated_urls else ""
        gen_str = "\n".join(generated_urls)

        r = [""] * 59
        r[0]  = cn_title
        r[1]  = en_title
        r[4]  = sku.get("variantName", "")
        r[5]  = sku.get("specName1", "")
        r[6]  = sku.get("specValue1", "")
        r[7]  = sku.get("specName2", "")
        r[8]  = sku.get("specValue2", "")
        r[9]  = first_generated
        r[10] = clean_price(shop.get("declarePrice") or sku.get("price", ""))
        r[12] = fmt_decimal(shop.get("length", ""))
        r[13] = fmt_decimal(shop.get("width", ""))
        r[14] = fmt_decimal(shop.get("height", ""))
        r[15] = fmt_decimal(shop.get("weight", ""))
        r[19] = gen_str
        r[20] = first_generated
        r[24] = clean_price(shop.get("retailPrice") or sku.get("price", ""))
        r[25] = shop.get("stock") or sku.get("stock", 0)
        r[26] = "9"
        r[27] = category_id
        r[28] = props_json
        r[29] = "[]"
        r[30] = sku.get("skcProps", "[]")
        r[31] = sku.get("skuProps", "[]")
        r[34] = shop.get("origin", "")
        r[37] = shop.get("skuClass") or "按件包装"
        r[38] = shop.get("skuClassQty") or "7"
        r[39] = shop.get("skuClassUnit") or "件"
        r[41] = "0"
        r[44] = "0"
        r[46] = "0"
        r[50] = video_url
        r[51] = shop.get("shipping", "")
        r[52] = shop.get("site", "")
        r[53] = shop.get("shopName", "")
        r[54] = sku.get("spuId", goods_id)
        r[55] = sku.get("skcId", "")
        r[56] = sku.get("skuId", "")
        r[57] = created_at
        r[58] = created_at

        for col, val in enumerate(r, 1):
            ws.cell(row=row_idx, column=col, value=val)
        row_idx += 1

    col_widths = [
        40, 40, 30, 15, 20, 15, 20, 15, 20, 50,
        12, 18, 8, 8, 8, 8, 12, 15, 50, 50,
        50, 12, 12, 50, 15, 8, 10, 10, 60, 10,
        50, 50, 12, 50, 20, 12, 30, 12, 10, 10,
        10, 10, 10, 10, 10, 10, 10, 10, 40, 12,
        50, 20, 20, 15, 20, 20, 20, 20, 20
    ]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    return output_path
